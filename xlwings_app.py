## import the required library
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import pyxlsb 
import os
import decimal
import xlwings as xw
import datetime
import re
import sys
from PyQt5.QtWidgets import *

#123
## 1.function1: Delete CPCT non-effectively date ---Start-------------------------------------------------------------------------------------------
def cpct_checker():
    ## Let user enters the effectively date, so get "text2" (Please refer to function1-1)
    cpctchecker_getdateForm()
    ## Let user chooses the CPCT files, so get "cpct_checkpath" (Please refer to function1-2)
    CPCT_checker_files() 

    ## -----below code is for testing
    # simulation = 'C:/Users/CheLily/OneDrive - HP Inc/Desktop/Quote/Quote Tool/test_iec/Quote Tool_20220303.xlsm' #for test
    # wb = xw.Book(simulation) #for test
    # base = wb.sheets['CPCT Checker'] #for test
    ## -----

    base = xw.Book.caller().sheets['CPCT Checker'] 
    base.range('D29', 'M87').clear_contents() 
    validated_date = text2
    all_paths_cpct = cpct_checkpath
    all_names_cpct = []

    #get the all of the CPCT file's name and store in all_names_cpct
    for i in cpct_checkpath:
        all_names_cpct.append(i.split('/')[-1])

    #If validated_date is not datetime format, change it to datetime format
    if isinstance(validated_date,datetime.datetime) == False:
        formatted_date1 = datetime.datetime.strptime(validated_date, "%m/%Y")
    else:
        formatted_date1 = validated_date

    row_no = 29
    if_tbd = 0
    for file, name in zip(all_paths_cpct, all_names_cpct) :
        print(name)
        revised_column_no = 0
        revised_column_no2 = 0
        #get sheet's name
        try:
            cpc_traker = load_workbook(file)
        except KeyError:
            continue
        sheet_names = cpc_traker.get_sheet_names()
        pricing_update = []
        spec_change = []
        #get the specific sheet's name
        for i in sheet_names:
            if 'CPC Tracker-pricing update' in i:
                pricing_update = i
            elif 'CPC Tracker-Pricing update' in i:
                pricing_update = i
            elif 'CPC Tracker-Pricing Update' in i:
                pricing_update = i
            elif 'CPC (GTK)Tracker-pricing update' in i:
                pricing_update = i                        
            if 'CPC Tracker-spec change' in i: 
                spec_change = i
            elif 'CPC Tracker-Spec change' in i:
                spec_change = i
            elif 'CPC Tracker-Spec Change' in i:
                spec_change = i
        #load data in tow different sheet
        worksheet1 = cpc_traker[pricing_update]
        if worksheet1['A3'].value == None:
            df1 = pd.read_excel(file, skiprows = 3, sheet_name = pricing_update)
            df2 = pd.read_excel(file, skiprows = 3, sheet_name = spec_change)
        else:
            df1 = pd.read_excel(file, skiprows = 2, sheet_name = pricing_update)
            df2 = pd.read_excel(file, skiprows = 2, sheet_name = spec_change) 




        #=============================================find the place that should be cleared(blank) in 'CPC Tracker-pricing update' sheet

        #find row and column of the cell that should be blank
        t = df1.loc[:,'Requestor':'Description'].head()
        revised_columns = t.iloc[:,1:-1].columns


        revised_index = []
        tbd_index = []
        for i, j, z in zip(df1['Effectivity Date'], df1['Request Date'], df1.index):
            #get the index is not effective

            #1.  Effectivity Date > validate date
            if isinstance(i,datetime.datetime) == True:                  
                if i > formatted_date1:
                    revised_index.append(z)
            #2.  Effectivity Date = Imme, and  Request Date > validate date
            elif i == 'Immed':
                if j > formatted_date1:
                    revised_index.append(z)
            #3.  TBD
            elif i == 'TBD':
                    tbd_index.append(z)         

        #get the index of the cell that should be blank
        revised_column_index = []
        for i in revised_columns.tolist():
            revised_column_index.append(df1.columns.get_loc(i))


        #=============================================find the place that should be cleared(blank) in 'CPC Tracker-spec change' sheet

        #find row and column of the cell that should be blank
        cleanedList = [x for x in df2.columns.tolist() if str(x) != 'nan']

        for i in cleanedList:
            if 'Description' in i:
                description = i
        t2 = df2.loc[:,'Requestor':description].head()
        revised_columns2 = t2.iloc[:,1:-1].columns

        revised_index2 = []
        tbd_index2 = []
        for i, j, z in zip(df2['Effectivity Date'], df2['Request Date'], df2.index):
            #get the index is not effective

            #1.  Effectivity Date > validate date
            if isinstance(i,datetime.datetime) == True:                  
                if i > formatted_date1:
                    revised_index2.append(z)
            #2.  Effectivity Date = Imme, and  Request Date > validate date
            elif i == 'Immed':
                if j > formatted_date1:
                    revised_index2.append(z)
            #3.  TBD
            elif i == 'TBD':
                    tbd_index2.append(z)


        #get the index of the cell that should be blank
        revised_column_index2 = []
        for i in revised_columns2.tolist():
            revised_column_index2.append(df2.columns.get_loc(i))            

        #=============================================update and color the cell that should be cleared in 'CPC Tracker-pricing' sheet     
        #color the cell that is not effective
        #app = xw.App(visible = False)
        #app.display_alerts = False
        #wb2 = xw.Book(file)
        app = xw.App(visible=False, add_book=False) # can't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb2 = app.books.open(file, update_links=False)

        sheet1 = wb2.sheets[pricing_update]

        for i in revised_column_index: #column
            for j in revised_index: #row
                if sheet1.range(j+4,i+1).value != np.nan and sheet1.range(j+4,i+1).value != None:
                    sheet1.range(j+4,i+1).color = (192, 0, 0)
                    sheet1.range(j+4,i+1).value = np.nan
                    revised_column_no +=1

        for i in revised_column_index: #column
            for j in tbd_index: #row
                if sheet1.range(j+4,i+1).value != np.nan and sheet1.range(j+4,i+1).value != None:
                    sheet1.range(j+4,i+1).color = (192, 0, 0)
                    sheet1.range(j+4,i+1).value = np.nan
                    if_tbd = 1
                    print_tbd = 1

        #=============================================update and color the cell that should be cleared in 'CPC Tracker-spec change' sheet  

        #color the cell that is not effective
        sheet2 = wb2.sheets[spec_change]
        for i in revised_column_index2: #column
            for j in revised_index2: #row
                if sheet2.range(j+4,i+1).value != np.nan and sheet2.range(j+4,i+1).value != None:
                    sheet2.range(j+4,i+1).color = (192, 0, 0)
                    sheet2.range(j+4,i+1).value = np.nan
                    revised_column_no2 +=1
        for i in revised_column_index2: #column
            for j in tbd_index2: #row
                if sheet2.range(j+4,i+1).value != np.nan and sheet2.range(j+4,i+1).value != None:
                    sheet2.range(j+4,i+1).color = (192, 0, 0)
                    sheet2.range(j+4,i+1).value = np.nan
                    if_tbd = 1
                    print_tbd = 1


        #=============================================show the number of cell and file name that were updated in the sheets 
        #     base_cpct = wb.sheets['CPCT Checker'] #for test
        base_cpct = xw.Book.caller().sheets['CPCT Checker'] 

        base_cpct.range(row_no,5).value = name
        base_cpct.range(row_no,8).value = revised_column_no   
        base_cpct.range(row_no,11).value = revised_column_no2
        if if_tbd == 1:
            base_cpct.range(row_no,12).value = "Please check your CPCT file: TBD with subcategory"
            print_tbd = 0
        row_no += 1

        # Save file: Create target Directory if it doesn't exist
        dirName = '/'.join(cpct_checkpath[0].split('/')[:-1])+'_revised'
        if not os.path.exists(dirName):
            os.mkdir(dirName)     
        wb2.save(dirName + '/' + name)
        wb2.close()
        app.kill()
        if_tbd = 0
    if if_tbd == 1:
        #When Effectivity Date = TBD, but the subcatexxx is not null, the Alert would pop out (Please refer to function1-4 )
        tbd_warn_code()
    base_cpct.range('B19').value = dirName
    #Let user user know the code is finished (pop up Done buttion.)(Please refer to function1-3)    
    finish_code()

## 1-1.function1-1: Let user enters the effectively date ---Start-------------------------------------------------------------------
class GetdateForm2(QWidget):
    def __init__(self, name = 'GetdateForm2'):
        super(GetdateForm2,self).__init__()
        self.setWindowTitle(name)
        self.cwd = os.getcwd() # get current code's file place
        self.resize(400,150)   # set the pop up widget's size
        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Enter Effective Month---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        
        
    def slot_btn_chooseFile1(self):
        global text2
        text2, okPressed = QInputDialog.getText(self, "Get text","Please enter Effective Month as in format like 01/2022", QLineEdit.Normal, "")
        if okPressed and text2 != '':
            self.close()
        
def cpctchecker_getdateForm():
    app = QApplication(sys.argv)
    getdateForm2 = GetdateForm2('CPCT Checker')
    getdateForm2.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...') 
## 1-1.function1-1: Let user enters the effectively date ---End-------------------------------------------------------------------

## 1-2.function1-2: Let user chooses the CPCT files  ---Start---------------------------------------------------------------------
class CPCT_checkerForm(QWidget):
    def __init__(self, name = 'CPCT_checkerForm'):
        super(CPCT_checkerForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose CPCT file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
    
    def slot_btn_chooseFile1(self):
        global cpct_checkpath
        cpct_checkpath, filetype = QFileDialog.getOpenFileNames(self,  
                                    "Choose CPCT file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if cpct_checkpath != "":
            self.close()

def CPCT_checker_files():
    app = QApplication(sys.argv)
    cpct_checkerForm = CPCT_checkerForm('CPCT Checker Form')
    cpct_checkerForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')


## 1-2.function1-2: Let user chooses the CPCT files ---End-------------------------------------------------------------------------

## 1-3.function1-3: Let user user know the code is finished (pop up Done buttion.) ---Start----------------------------------------
class CloseForm(QWidget):
    def __init__(self, name = 'CloseForm'):
        super(CloseForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(200,100)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_done")  
        self.btn_done.setText("Done")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def finish_code():
    app = QApplication(sys.argv)
    closeForm = CloseForm('Quote Validation')
    closeForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 1-3.function1-3: Let user user know the code is finished (pop up Done buttion.) ---Done-----------------------------------------

## 1-4.function1-4: When Effectivity Date = TBD, but the subcatexxx is not null, the Alert would pop out  ---Start----------------------------------------
class TBD_Warn_Form(QWidget):
    def __init__(self, name = 'TBD_Warn_Form'):
        super(TBD_Warn_Form,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_tbd_warn")  
        self.btn_done.setText("Please check your CPCT file: TBD with subcategory")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def tbd_warn_code():
    app = QApplication(sys.argv)
    tbd_warn_Form = TBD_Warn_Form('Quote Validation')
    tbd_warn_Form.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')

## 1-4.function1-4: When Effectivity Date = TBD, but the subcatexxx is not null, the Alert would pop out ---Done-----------------------------------------

## 1.function1: Delete CPCT non-effectively date ---Done-----------------------------------------------------------------------------------------------------------------------------------------------------






## 2.function2: Quote Validation ---Start-----------------------------------------------------------------------------------------------------------------------------------------------------------------
def quote_validation():
    import xlwings as xw

    ## -----below quote is for testing
    # simulation = 'C:/Users/CheLily/OneDrive - HP Inc/Desktop/Quote/Quote Validation-IEC/test-quote tool/Quote Tool_20220308_v7.xlsm' #for test
    # wb = xw.Book(simulation) #for test
    # base = wb.sheets['Quote Validation'] #for test
    ## -----
    base = xw.Book.caller().sheets['Quote Validation'] 
    ## Let user enters the quote file, so get "QuoteName_choose" (Please refer to function2-1)
    get_quoteform()
    ## Let user chooses the CPCT, so get "fileName_busa" (Please refer to function2-2)
    Consolidation_BUSA()
    ## Let user enters the Sell Price&CKIT file, so get "fileName_choose1", "fileName_choose2", "fileName_choose3" (Please refer to function2-3)
    choose_files2()
    #### Get quote from Progarm matrix
    df = pd.DataFrame(columns=['HP P/N', 'Type', 'Current Price']) # CKIT 
    df2 = pd.DataFrame(columns=['HP P/N', 'Description', 'Total Base Unit Cost excluded B/S', 'Platform']) # BUSA
    df3 = pd.DataFrame(columns=['HP P/N', 'Description', 'Category', 'Current Price']) # Master Data

    for paths_quote in QuoteName_choose:
        base.range('E21').value = paths_quote
        #get quote_program_matrix sheet's name
        if 'xlsb' in paths_quote:
            quote_program_matrix = pd.ExcelFile(paths_quote, engine='pyxlsb')
        else:
            quote_program_matrix = pd.ExcelFile(paths_quote)

        sheet_names = quote_program_matrix.sheet_names 

        sheet_name_changelog = []
        sheet_name_master_data = []
        sheet_name_ckit = []
        sheet_name_busa = []
        for i in sheet_names:
            if 'Change Log' in i:
                sheet_name_changelog = i
            if 'Change log' in i:
                sheet_name_changelog = i
            if 'Master Data' in i: 
                sheet_name_master_data = i
            if 'Master data' in i: 
                sheet_name_master_data = i                
            if 'CKIT' in i: 
                sheet_name_ckit = i
            if 'BUSA' in i: 
                sheet_name_busa = i            
            if 'BU SA' in i: 
                sheet_name_busa = i 

        #get platform & ODM's name
        import xlwings as xw
        app = xw.App(visible=False, add_book=False) # can't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_quote, update_links=False, read_only=True, ignore_read_only_recommended=True)

        sheet_change_log = wb.sheets[sheet_name_changelog]    
        platform_name = sheet_change_log.range('B2').value
        
        odm_name = sheet_change_log.range('B3').value
        quote_file_name = paths_quote.split("/")[-1]

        #load the data
        if 'xlsb' in paths_quote:
            df_quote_masterdata = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_master_data)
            df_quote_ckit = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_ckit)
            df_quote_busa = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_busa)

        else: 
            df_quote_masterdata = pd.read_excel(paths_quote, skiprows = 0, sheet_name = sheet_name_master_data)
            df_quote_ckit = pd.read_excel(paths_quote, skiprows = 0, sheet_name = sheet_name_ckit)
            df_quote_busa = pd.read_excel(paths_quote, skiprows = 0, sheet_name = sheet_name_busa)

        # get "Master Data" & "CKIT" & "BU SA" columns' name-------------------------------------------------------------------------------------
        df_quote_masterdata_column_name = df_quote_masterdata.columns.tolist()
        for i in df_quote_masterdata_column_name:
            if 'HP' in i:
                effective_column_name1_1 = i
            if 'Current' in i:
                effective_column_name1_2 = i
        df_quote_masterdata = df_quote_masterdata.reindex(columns=[effective_column_name1_1, 'Description', effective_column_name1_2, 'Category'])
        # df_quote_masterdata = df_quote_masterdata.dropna(how='any')

        df_quote_ckit_column_name = df_quote_ckit.columns.tolist()
        for i in df_quote_ckit_column_name:
            if 'Current' in i:
                effective_column_name2 = i
        df_quote_ckit = df_quote_ckit.reindex(columns=['HP Part #', 'Type', effective_column_name2])
        # df_quote_ckit = df_quote_ckit.dropna(how='any')

        df_quote_busa_column_name = df_quote_busa.columns.tolist()
        for i in  df_quote_busa_column_name:
            if 'HP' in i:
                effective_column_name3 = i
            elif 'SA' in i:
                effective_column_name3 = i
        df_quote_busa = df_quote_busa.reindex(columns=[effective_column_name3, 'Description', 'Total Base Unit Cost excluded B/S'])
        # df_quote_busa = df_quote_busa.dropna(how='any')

        # get "Master Data" & "CKIT" & "BU SA" columns' info----------------------------------------------------------------------------------------------------
        df_quote_masterdata.rename(columns={effective_column_name1_1:'HP P/N', effective_column_name1_2:'Current Price'}, inplace=True)
        df_quote_masterdata['Platform'] = platform_name
        df_quote_masterdata['ODM'] = odm_name
        df_quote_masterdata['Quote File Name'] = quote_file_name
    #     df_quote_avsummary = df_quote_avsummary[~df_quote_avsummary['AV'].isnull()] #delete row which SKU is null
        df3 = df3.append(df_quote_masterdata) 


        df_quote_ckit.rename(columns={'HP Part #':'HP P/N', effective_column_name2:'Current Price'}, inplace=True)
        df_quote_ckit['Platform'] = platform_name
        df_quote_ckit['ODM'] = odm_name
        df_quote_ckit['Quote File Name'] = quote_file_name
    #     df_quote_avsummary = df_quote_avsummary[~df_quote_avsummary['AV'].isnull()] #delete row which SKU is null
        df = df.append(df_quote_ckit) 

        df_quote_busa.rename(columns={effective_column_name3:'HP P/N'}, inplace=True)
        df_quote_busa['Platform'] = platform_name
        df_quote_busa['ODM'] = odm_name
        df_quote_busa['Quote File Name'] = quote_file_name
    #     df_quote_avsummary = df_quote_avsummary[~df_quote_avsummary['AV'].isnull()] #delete row which SKU is null
        df2 = df2.append(df_quote_busa) 
        print(paths_quote)
        wb.close() # close file
        app.quit() # close app



    #### Get BUSA & OP quote from CPCT's BUSA sheet-----------------------------------------------------------------------------------------
    #Add OP Quote -- Lily 2022/05/18
    busa_base = pd.DataFrame(columns=['SA \nLevel 3', 'Description', 'Total Base Unit Cost excluded B/S', 'Path'])
    op_base = pd.DataFrame(columns=['SA PartNumber', 'SA Description', 'Total Cost', 'Path'])
    sheet_op_missing = []
    for paths_cpct in fileName_busa:
        base.range('E22').value = paths_cpct
        #open file
        import xlwings as xw
        app = xw.App(visible=False, add_book=False) # don't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_cpct, update_links=False, read_only=True, ignore_read_only_recommended=True)

        if 'xlsb' in paths_cpct:
            cpct = pd.ExcelFile(paths_cpct, engine='pyxlsb')
        else:
            cpct = pd.ExcelFile(paths_cpct)            
        sheet_names = cpct.sheet_names 

        sheet_name_busa = []
        sheet_name_op = []
        for i in sheet_names:
            if 'BUSA' in i: 
                sheet_name_busa = i            
            if 'BU SA' in i: 
                sheet_name_busa = i 
            if 'OptionSA_SUM' in i: 
                sheet_name_op = i
                
        #load the data
        #revised by Lily (lily.chen1@hp.com) to prevent 'OptionSA_SUM' sheet missing on 2022/06/10
        if 'xlsb' in paths_cpct:
            df_busa = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_busa)
        else: 
            df_busa = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = sheet_name_busa)
        
        
        if sheet_name_op != []:
            if 'xlsb' in paths_cpct:
                df_op = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_op)
            else: 
                df_op = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = sheet_name_op)
        else:
            df_op = pd.DataFrame(columns=['SA PartNumber', 'SA Description', 'Total Cost', 'Program Matrix']) 
            sheet_op_missing.append(paths_cpct)
        
        df_busa['Path'] = paths_cpct
        df_op['Path'] = paths_cpct
        # append
        busa_base = busa_base.append(df_busa) 
        op_base = op_base.append(df_op, ignore_index = True) 

        wb.close() # close file
        app.quit() # close app

    #### Compare the quote from both program matrix and CPCT's BUSA---------------------------------------------------------------------------
    # xlookup function that is similar to vlookup in excel (Please refer to function2-4)
    df2['HP Price'] = df2['HP P/N'].apply(xlookup, args = (busa_base['SA \nLevel 3'], busa_base['Total Base Unit Cost excluded B/S']))
    df_list = []
    for i in df2['HP Price']:
        if isinstance(i,str):
            df_list.append(np.nan)
        elif isinstance(i,datetime.datetime):
            df_list.append(np.nan)
        elif isinstance(i,datetime.time):
            df_list.append(np.nan)        
        else:
            df_list.append(i)
    df2['HP Price']  = df_list

    df_list2 = []
    for i in df2['Total Base Unit Cost excluded B/S']:
        if isinstance(i,str):
            df_list2.append(np.nan)
        elif isinstance(i,datetime.datetime):
            df_list2.append(np.nan)
        elif isinstance(i,datetime.time):
            df_list2.append(np.nan)
        else:
            df_list2.append(i)
    df2['Total Base Unit Cost excluded B/S']  = df_list2         
    df2['Delta'] = df2['Total Base Unit Cost excluded B/S'] - df2['HP Price']

    #### Get sell price from Sell price file
    if 'xlsb' in fileName_choose1:
        sell_base = pd.read_excel(fileName_choose1, engine='pyxlsb', skiprows = 0, sheet_name = 'Master Data')
    else: 
        sell_base = pd.read_excel(fileName_choose1, skiprows = 0, sheet_name = 'Master Data')
    sell_base['Path'] = fileName_choose1

    #### Compare the quote from both program matrix and Sell price file--------------------------------------------------------------------------
    df3['HP Price'] = df3['HP P/N'].apply(xlookup, args = (sell_base['HP P/N'], sell_base['Current Price']))
    df_list = []
    for i in df3['HP Price']:
        if isinstance(i,str):
            df_list.append(np.nan)
        elif isinstance(i,datetime.datetime):
            df_list.append(np.nan)
        elif isinstance(i,datetime.time):
            df_list.append(np.nan)        
        else:
            df_list.append(i)
    df3['HP Price']  = df_list

    df_list2 = []
    for i in df3['Current Price']:
        if isinstance(i,str):
            df_list2.append(np.nan)
        elif isinstance(i,datetime.datetime):
            df_list2.append(np.nan)
        elif isinstance(i,datetime.time):
            df_list2.append(np.nan)
        else:
            df_list2.append(i)
    df3['Current Price']  = df_list2  
    df3['Delta'] = df3['Current Price'] - df3['HP Price']

    #### Get ckit price from CKIT file  ----------------------------------------------------------------------------------------------------
    if 'xlsb' in fileName_choose2:
        df_ckit_nb1 = pd.read_excel(fileName_choose2, engine='pyxlsb', skiprows = 0, sheet_name = 'Doc Kit SKU Summary')
        df_ckit_nb2 = pd.read_excel(fileName_choose2, engine='pyxlsb', skiprows = 0, sheet_name = 'Doc KIT SKU summary for HP')
    else:
        df_ckit_nb1 = pd.read_excel(fileName_choose2, skiprows = 0, sheet_name = 'Doc Kit SKU Summary')
        df_ckit_nb2 = pd.read_excel(fileName_choose2, skiprows = 0, sheet_name = 'Doc KIT SKU summary for HP')

    df_ckit_me = pd.DataFrame(columns = ['HP P/N', 'Description', 'Current Price']) 
    ckit_file = load_workbook(fileName_choose3)
    skit_sheet_names = ckit_file.get_sheet_names()
    for i, j in zip(ckit_file, ckit_file.sheetnames):
        if i.sheet_state == "visible":
            if 'xlsb' in fileName_choose3:
                df_ckit_me1 = pd.read_excel(fileName_choose3, engine='pyxlsb', skiprows = 0, sheet_name = j)
                df_ckit_me = pd.concat((df_ckit_me, df_ckit_me1[['HP P/N', 'Description', 'Current Price']]), axis=0, ignore_index=True)
            else:
                df_ckit_me1 = pd.read_excel(fileName_choose3, skiprows = 0, sheet_name = j) 
                df_ckit_me = pd.concat((df_ckit_me, df_ckit_me1[['HP P/N', 'Description', 'Current Price']]), axis=0, ignore_index=True)

    df_ckit_nb = pd.concat([df_ckit_nb1[['HP P/N', 'Description', 'Current Price']], df_ckit_nb2[['HP P/N', 'Description', 'Current Price']]],axis=0, ignore_index=True)
    ckit_base = pd.concat([df_ckit_nb, df_ckit_me],axis=0, ignore_index=True)    
    ckit_base['Path'] = fileName_choose2

    #clean ckit data - deal with the issue caused by 'Merge cells'--------------------------------------------------------------------------
    ckit_base['HP P/N'] = ckit_base['HP P/N'].str.replace('\n','')
    ckit_base['Current Price'].fillna(method='ffill', inplace=True)

    #### Compare the quote from both program matrix and Sell price file--------------------------------------------------------------------------
    df['HP Price'] = df['HP P/N'].apply(xlookup, args = (ckit_base['HP P/N'], ckit_base["Current Price"]))


    df_list = []
    for i in df['HP Price']:
        if isinstance(i,str):
            df_list.append(np.nan)
        elif isinstance(i,datetime.datetime):
            df_list.append(np.nan)
        elif isinstance(i,datetime.time):
            df_list.append(np.nan)        
        else:
            df_list.append(i)
    df['HP Price']  = df_list

    df_list2 = []
    for i in df['Current Price']:
        if isinstance(i,str):
            df_list2.append(np.nan)
        elif isinstance(i,datetime.datetime):
            df_list2.append(np.nan)
        elif isinstance(i,datetime.time):
            df_list2.append(np.nan)
        else:
            df_list2.append(i)
    df['Current Price']  = df_list2  

    df['Delta'] = df['Current Price'] - df['HP Price']

    # Consolidate HP price----------------------------------------------------------------------------------------------------------------------------------------------------
    ckit_base['Category'] = 'CKIT'
    busa_base['Category'] = 'BU SA'
    op_base['Category'] = 'OP'
    sell_base['Category'] = 'Sell Price'
    busa_base = busa_base[['SA \nLevel 3', 'Description', 'Total Base Unit Cost excluded B/S', 'Category', 'Path']]
    op_base = op_base[['SA PartNumber', 'SA Description', 'Total Cost', 'Category', 'Path']]
    busa_base.set_axis(['HP P/N', 'Description', 'Current Price', 'Category', 'Path'],axis='columns', inplace=True)
    op_base.set_axis(['HP P/N', 'Description', 'Current Price', 'Category', 'Path'],axis='columns', inplace=True)
    HP_price_base = pd.concat([ckit_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']], busa_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']], sell_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']], op_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']]], axis=0, ignore_index=True)    
   
    #### save file----------------------------------------------------------------------------------------------------------------------------------------------------
        # Create target Directory if it doesn't exist
    dirName = '/'.join(QuoteName_choose[0].split('/')[:-2])+'/Quote_Validation'
    if not os.path.exists(dirName):
        os.mkdir(dirName) 

        #create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(dirName+'/Quote_Validation_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx', engine='xlsxwriter')

        #write each DataFrame to a specific sheet
    df.to_excel(writer, sheet_name='CKIT', index=False)
    df2.to_excel(writer, sheet_name='BU SA', index=False)
    df3.to_excel(writer, sheet_name='Sell Price', index=False)
    HP_price_base.to_excel(writer, sheet_name='HP Price', index=False)
    #format header
        # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['CKIT']
    worksheet2 = writer.sheets['BU SA']
    worksheet3 = writer.sheets['Sell Price']
    worksheet4 = writer.sheets['HP Price']

        # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#8EA9DB'})

    worksheet.set_column("A:H", 25)
    worksheet2.set_column("A:I", 25)
    worksheet3.set_column("A:I", 25)
    worksheet4.set_column("A:E", 25)

        # Write the column headers with the defined format.
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for col_num, value in enumerate(df2.columns.values):
        worksheet2.write(0, col_num, value, header_format)
    for col_num, value in enumerate(df3.columns.values):
        worksheet3.write(0, col_num, value, header_format)
    for col_num, value in enumerate(HP_price_base.columns.values):
        worksheet4.write(0, col_num, value, header_format)
        
        #close the Pandas Excel writer and output the Excel file--------------------------------------------------------------------------
    writer.save()
    writer.close()
    base.range('B19').value = dirName+'/Quote_Validation_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx'
    base.range('C29', 'Q200').clear_contents()
    base.range('C29').value = pd.DataFrame(QuoteName_choose)
    #revised by Lily (lily.chen1@hp.com) to prevent 'OptionSA_SUM' sheet missing on 2022/06/10
    #print  CPCT that op sheet missing in the Quote tool
    if len(sheet_op_missing) > 0:
        base.range('S29').value = pd.DataFrame(sheet_op_missing)
        #pop up warning
        MissingOP()
    #Let user user know the code is finished (pop up Done buttion.)(Please refer to function1-3)
    finish_code()

## 2-1.function2-1: Let user chooses the Quote files ---Start-------------------------------------------------------------------------
class Get_QuoteForm(QWidget):
    def __init__(self, name = 'Get_QuoteForm'):
        super(Get_QuoteForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose Quote file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
    
    def slot_btn_chooseFile1(self):
        global QuoteName_choose
        QuoteName_choose, filetype = QFileDialog.getOpenFileNames(self,  
                                    "Choose Quote file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if QuoteName_choose != "":
            self.close()

def get_quoteform():
    app = QApplication(sys.argv)
    get_quoteform = Get_QuoteForm('Get_QuoteForm')
    get_quoteform.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 2-1.function2-1: Let user chooses the Quote files ---End-------------------------------------------------------------------------

## 2-2.function2-2: Let user chooses the CPCT files ---Start-------------------------------------------------------------------------
class ConsolidationForm_BUSA(QWidget):
    def __init__(self, name = 'ConsolidationForm_BUSA'):
        super(ConsolidationForm_BUSA,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose CPCT file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
    
    def slot_btn_chooseFile1(self):
        global fileName_busa
        fileName_busa, filetype = QFileDialog.getOpenFileNames(self,  
                                    "Choose CPCT file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_busa != "":
            self.close()
            
def Consolidation_BUSA():
    app = QApplication(sys.argv)
    consolidationForm_BUSA = ConsolidationForm_BUSA('Quote Consolidation_BUSA')
    consolidationForm_BUSA.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 2-2.function2-2: Let user chooses the CPCT files ---End-------------------------------------------------------------------------

## 2-3.function2-3: Let user enters the Sell Price&CKIT files---Start-------------------------------------------------------------------------
# Get ODM & CKIT file path
class MainForm2(QWidget):
    def __init__(self, name = 'MainForm2'):
        super(MainForm2,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 2
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose Sell Price file's Path---]")

        # btn 3
        self.btn_chooseFile2 = QPushButton(self)  
        self.btn_chooseFile2.setObjectName("btn_chooseFile")  
        self.btn_chooseFile2.setText("[---Choose CKIT-NB file's Path---]")

        # btn 3
        self.btn_chooseFile3 = QPushButton(self)  
        self.btn_chooseFile3.setObjectName("btn_chooseFile")  
        self.btn_chooseFile3.setText("[---Choose CKIT-Media file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)
        layout.addWidget(self.btn_chooseFile2)
        layout.addWidget(self.btn_chooseFile3)
        
        self.setLayout(layout)


        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        self.btn_chooseFile2.clicked.connect(self.slot_btn_chooseFile2)
        self.btn_chooseFile3.clicked.connect(self.slot_btn_chooseFile3)
        
    def slot_btn_chooseFile1(self):
        global fileName_choose1
        fileName_choose1, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose Sell Price file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

#         if fileName_choose1 != "":
#             self.close()
#         elif fileName_choose1 == "":
#             return 
#         base = xw.Book.caller().sheets['Main']
#         base.range('F13').value = fileName_choose


    def slot_btn_chooseFile2(self):
        global fileName_choose2
        fileName_choose2, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose CKIT-NB file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")  

#         if fileName_choose2 != "":
#             self.close()
#         elif fileName_choose2 == "":
#             return 

#         base = xw.Book.caller().sheets['Main']
#         base.range('F17').value = fileName_choose

    def slot_btn_chooseFile3(self):
        global fileName_choose3
        fileName_choose3, filetype = QFileDialog.getOpenFileName(self,  
                                    "Choose CKIT-Media file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")  

        if fileName_choose3 != "":
            self.close()
        elif fileName_choose3 == "":
            return 
            

#         base = xw.Book.caller().sheets['Main']
#         base.range('F17').value = fileName_choose
 
def choose_files2():
    app = QApplication(sys.argv)
    mainForm2 = MainForm2('Quote Validation')
    mainForm2.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')

## 2-3.function2-3: Let user enters the Sell Price&CKIT files---End-------------------------------------------------------------------------


## 2-4.function2-4: Create a function that is similar to vlookup in excel---Start-------------------------------------------------------------------------
def xlookup(lookup_value, lookup_array, return_array, if_not_found:str = ''):
    match_value = return_array.loc[lookup_array == lookup_value]
    if match_value.empty:
#         return f'"{lookup_value}" not found!' if if_not_found == '' else if_not_found
        return np.nan if if_not_found == '' else if_not_found

    else:
        return match_value.tolist()[0]
## 2-4.function2-4: Create a function that is similar to vlookup in excel---End-------------------------------------------------------------------------

#revised by Lily (lily.chen1@hp.com) to prevent 'OptionSA_SUM' sheet missing on 2022/06/10
## 2-5.function2-5: Let user know "OP SUM' sheet is missing in CPCT ---Start----------------------------------------
class MissingOPForm(QWidget):
    def __init__(self, name = 'MissingOPForm'):
        super(MissingOPForm,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_done")  
        self.btn_done.setText("Please check your CPCT: OP SUM sheet is missing")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def MissingOP():
    app = QApplication(sys.argv)
    missingOPForm = MissingOPForm('Quote Validation')
    missingOPForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 2-5.function2-5: Let user know "OP SUM' sheet is missing in CPCT ---Start----------------------------------------

## 2.function2: Quote Validation ---End-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------




            
        
        
        
## 3.function3: Quote Consolidation ---Start-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def quote_consolidation():
    #Let user enters effectively date(Please reder to function3-2)
    consolidation_getdateForm()    
    #Let user chooses the quote file to consolidate(Please reder to function3-1)
    app = QApplication(sys.argv)
    consolidationForm = ConsolidationForm('Quote Consolidation')
    consolidationForm.show()
    
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
    
    
    new_effective_day = 'Effective '+ text
    df = pd.DataFrame(columns=['AV', 'Description', new_effective_day, 'Platform', 'ODM', 'Quote File Name'])
    df2 = pd.DataFrame(columns=['SKU', 'Description', new_effective_day, 'Platform', 'ODM', 'Quote File Name'])
    test_platform = []
    for paths_quote in fileName_choose:
          
        test_platform.append(paths_quote)
#         print(paths_quote)
         #get quote_program_matrix sheet's name
        if 'xlsb' in paths_quote:
            quote_program_matrix = pd.ExcelFile(paths_quote, engine='pyxlsb')
        else:
            quote_program_matrix = pd.ExcelFile(paths_quote)

        sheet_names = quote_program_matrix.sheet_names 

#         sheet_name_change_log = []
#         sheet_name_avsummary = []
#         sheet_name_skusummary = []
#         for i in sheet_names:
#             if 'Change Log' in i:
#                 sheet_name_changelog = i
#             if 'Change log' in i:
#                 sheet_name_changelog = i
#             if 'AV Summary' in i: 
#                 sheet_name_avsummary = i
#             if 'AV summary' in i: 
#                 sheet_name_avsummary = i
#             if 'SKU Summary' in i: 
#                 sheet_name_skusummary = i
#             if 'SKU summary' in i: 
#                 sheet_name_skusummary = i

        #get platform & ODM's name
        app = xw.App(visible=False, add_book=False) # can't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_quote, update_links=False, read_only=True, ignore_read_only_recommended=True)

        sheet_change_log = wb.sheets['Change Log']    
        platform_name = sheet_change_log.range('B2').value
        odm_name = sheet_change_log.range('B3').value
        quote_file_name = paths_quote.split("/")[-1]

        #load the data-----------------------------------------------------------------------------------------------------------
        if 'AV Summary' in sheet_names:
            if 'xlsb' in paths_quote:
                df_quote_avsummary = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = 'AV Summary')
            else: 
                df_quote_avsummary = pd.read_excel(paths_quote, skiprows = 0, sheet_name = 'AV Summary')
        else:
            df_quote_avsummary = pd.DataFrame(columns=['AV', 'Description', 'Current Month'])
        if 'SKU Summary' in sheet_names:
            if 'xlsb' in paths_quote:
                df_quote_skusummary = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = 'SKU Summary')
            else:
                df_quote_skusummary = pd.read_excel(paths_quote, skiprows = 0, sheet_name = 'SKU Summary')
        else:
            df_quote_skusummary = pd.DataFrame(columns=['SKU', 'Description', 'Current Month'])
            
        # get "AV Summary" & "SKU Summary" columns' name--------------------------------------------------------------------------
        df_quote_avsummary_column_name = df_quote_avsummary.columns.tolist()
        df_quote_avsummary_column_name = [str(int) for int in df_quote_avsummary_column_name]
        effective_column_name = 'Current Month'
        for i in df_quote_avsummary_column_name:
            if 'Effective' in i:
                effective_column_name = i
            elif 'Current' in i:  
                effective_column_name = i
        df_quote_avsummary = df_quote_avsummary.reindex(columns=['AV', 'Description', effective_column_name])

        df_quote_skusummary_column_name = df_quote_skusummary.columns.tolist()
        effective_column_name2 = 'Current Month'
        for i in df_quote_skusummary_column_name:
            if 'Effective' in i:
                effective_column_name2 = i
            elif 'Current' in i:  
                effective_column_name2 = i               
        df_quote_skusummary = df_quote_skusummary.reindex(columns=['SKU', 'Description', effective_column_name2])

        # get "AV Summary" & "SKU Summary" columns' info------------------------------------------------------------------------
        df_quote_avsummary.rename(columns={effective_column_name:new_effective_day}, inplace=True)
        df_quote_avsummary['Platform'] = platform_name
        df_quote_avsummary['ODM'] = odm_name
        df_quote_avsummary['Quote File Name'] = quote_file_name
        df_quote_avsummary = df_quote_avsummary[~df_quote_avsummary['AV'].isnull()] #delete row which Description is null

        # update comment----------------------------------------------------------------------------------------------------------
        df_quote_avsummary['Comment'] = ''
        df_list = []
        for i in df_quote_avsummary[new_effective_day]:
            if isinstance(i,str):
                df_list.append(np.nan)       
            else:
                df_list.append(i)
        df_quote_avsummary[new_effective_day] = df_list
        # df_quote_avsummary.loc[df_quote_avsummary[new_effective_day] == '0x2a', new_effective_day] = np.nan
        df_quote_avsummary.loc[df_quote_avsummary[new_effective_day] == 0, 'Comment'] = 'AV Cost Not Updated'
        df_quote_avsummary.loc[df_quote_avsummary[new_effective_day].isnull(), new_effective_day] = '#N/A'
        df_quote_av_not_duplicate = df_quote_avsummary[~df_quote_avsummary.duplicated(subset=['AV',new_effective_day])]
        df_quote_av_same = df_quote_av_not_duplicate[df_quote_av_not_duplicate.duplicated(subset='AV')]['AV'].tolist()
        df_quote_avsummary.loc[df_quote_avsummary['AV'].isin(df_quote_av_same), 'Comment'] = 'Same AV With Different Cost'
        df = df.append(df_quote_avsummary) 
        


        df_quote_skusummary.rename(columns={effective_column_name2:new_effective_day}, inplace=True)
        df_quote_skusummary['Platform'] = platform_name
        df_quote_skusummary['ODM'] = odm_name
        df_quote_skusummary['Quote File Name'] = quote_file_name
        df_quote_skusummary = df_quote_skusummary[~df_quote_skusummary['SKU'].isnull()] #delete row which Description is null


        # update comment-------------------------------------------------------------------------------------------------------------
        df_quote_skusummary['Comment'] = ''
        df_list2 = []
        for i in df_quote_skusummary[new_effective_day]:
            if isinstance(i,str):
                df_list2.append(np.nan)       
            else:
                df_list2.append(i)
        df_quote_skusummary[new_effective_day] = df_list2
        # df_quote_skusummary.loc[df_quote_skusummary[new_effective_day] == '0x2a', new_effective_day] = np.nan
        df_quote_skusummary.loc[df_quote_skusummary[new_effective_day] == 0, 'Comment'] = 'SKU Cost Not Updated'
        df_quote_skusummary.loc[df_quote_skusummary[new_effective_day].isnull(), new_effective_day] = '#N/A'
        df_quote_sku_same = df_quote_skusummary[df_quote_skusummary.duplicated(subset='SKU')]
        df_quote_price_diff2 = df_quote_sku_same[~df_quote_sku_same.duplicated(subset=new_effective_day)]['SKU'].tolist()
        df_quote_skusummary.loc[df_quote_skusummary['SKU'].isin(df_quote_price_diff2), 'Comment'] = 'Same SKU With Different Cost'
        df2 = df2.append(df_quote_skusummary)
#        print(paths_quote)
        wb.close() # close file
        app.quit() # close app

    #save file-------------------------------------------------------------------------------------------------------------------------------------
        # Create target Directory if it doesn't exist
    dirName = '/'.join(fileName_choose[0].split('/')[:-2])+'/Quote_Consolidation_' + text 
    if not os.path.exists(dirName):
        os.mkdir(dirName) 

        #create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(dirName+'/Quote_Consolidation_'+text+'_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx', engine='xlsxwriter')

        #write each DataFrame to a specific sheet
    df.to_excel(writer, sheet_name='AV Summary', index=False)
    df2.to_excel(writer, sheet_name='SKU Summary', index=False)
    #format header
        # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['AV Summary']
    worksheet2 = writer.sheets['SKU Summary']

        # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#8EA9DB'})

    worksheet.set_column("A:G", 25)
    worksheet2.set_column("A:G", 25)

        # Write the column headers with the defined format.
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for col_num, value in enumerate(df2.columns.values):
        worksheet2.write(0, col_num, value, header_format)


        #close the Pandas Excel writer and output the Excel file
    writer.save()
    writer.close()
    #print save file's path
    base = xw.Book.caller().sheets['Quote Consolidation'] 
    base.range('B19').value = dirName+'/Quote_Consolidation_'+text+'_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx'
    df3 = pd.DataFrame(test_platform)
    base.range('C29', 'O311').clear_contents()
    base.range('C29').value = df3
    finish_code()



## 3-1.function3-1: Let user choose the quote file to consolidate ---Start------------------------------------------------------------------
class ConsolidationForm(QWidget):
    def __init__(self, name = 'ConsolidationForm'):
        super(ConsolidationForm,self).__init__()
        self.setWindowTitle(name)
        self.cwd = os.getcwd() # get current code's file place
        self.resize(400,150)   # set the pop up widget's size

        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose Quote file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        self.close()
        
    def slot_btn_chooseFile1(self):
        global fileName_choose
        fileName_choose, filetype = QFileDialog.getOpenFileNames(self,  
                                    "Choose Quote file's Path",  
                                    self.cwd, # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_choose != "":
            self.close()
## 3-1.function3-1: Let user choose the quote file to consolidate ---End------------------------------------------------------------------

## 3-2.function3-2: Let user enter effectively date ---Start------------------------------------------------------------------
class GetdateForm(QWidget):
    def __init__(self, name = 'GetdateForm'):
        super(GetdateForm,self).__init__()
        self.setWindowTitle(name)
        self.cwd = os.getcwd() # get current code's file place
        self.resize(400,150)   # set the pop up widget's size
        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Enter Effective Month---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
        
        
    def slot_btn_chooseFile1(self):
        global text
        text, okPressed = QInputDialog.getText(self, "Get text","Please enter Effective Month as in format like Sep-2021", QLineEdit.Normal, "")
        if okPressed and text != '':
            self.close()
        
def consolidation_getdateForm():
    app = QApplication(sys.argv)
    getdateForm = GetdateForm('Quote Consolidation')
    getdateForm.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')  

## 3-2.function3-2: Let user enter effectively date ---End------------------------------------------------------------------



## 3.function3: Quote Consolidation ---End--------------------------------------------------------------------------------------------------------------------------------------------------------------------


        
## 4.function4: CPCT BU SA & Option Consolidation ---Start-----------------------------------------------------------------------------------------------------------------------------------------------------------------
   ## (Same as compal - "2.function2: CPCT BU SA & Option Consolidation")
def cpct_consolidation():
    import xlwings as xw

    ## -----below code is for testing
    #     simulation = 'C:/Users/CheLily/OneDrive - HP Inc/Desktop/Quote/Quote Tool/test_compal/Quote Tool_20220425_v12.xlsm' #for test
    #     wb = xw.Book(simulation) #for test
    #     base = wb.sheets['CPCT Consolidation'] #for test
    ## -----    

    base = xw.Book.caller().sheets['CPCT Consolidation'] 
    ## Let user chooses the CPCT, so get "fileName_busa" (Please refer to function4-1)
    Consolidation_BUSA()

    #### Get BUSA & OP quote from CPCT's BUSA sheet
    busa_base = pd.DataFrame(columns=['SA \nLevel 3', 'Description', 'Total Base Unit Cost excluded B/S', 'Platform', 'Path'])
    op_base = pd.DataFrame(columns=['SA PartNumber', 'SA Description', 'Total Cost', 'Path'])
    sheet_op_missing = []
    for paths_cpct in fileName_busa:
        base.range('E22').value = paths_cpct
        #open file
        import xlwings as xw
        app = xw.App(visible=False, add_book=False) # don't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_cpct, update_links=False, read_only=True, ignore_read_only_recommended=True)

        if 'xlsb' in paths_cpct:
            cpct = pd.ExcelFile(paths_cpct, engine='pyxlsb')
        else:
            cpct = pd.ExcelFile(paths_cpct)            
        sheet_names = cpct.sheet_names 

        sheet_name_busa = []
        sheet_name_op = []
        for i in sheet_names:
            if 'BUSA' in i: 
                sheet_name_busa = i            
            if 'BU SA' in i: 
                sheet_name_busa = i 
            if 'OptionSA_SUM' in i: 
                sheet_name_op = i            

        #load the BU SA data
        if 'xlsb' in paths_cpct:
            df_busa = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_busa)
        else: 
            df_busa = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = sheet_name_busa)
        df_busa['Path'] = paths_cpct
        # append
        busa_base = busa_base.append(df_busa) 
        busa_base = busa_base[['SA \nLevel 3', 'Description', 'Total Base Unit Cost excluded B/S', 'Platform', 'Path']]
        #load OptionSA_SUM quote
        #revised by Lily (lily.chen1@hp.com) to prevent 'OptionSA_SUM' sheet missing on 2022/06/10
        if sheet_name_op != []:
            if 'xlsb' in paths_cpct:
                df_op = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_op)
            else: 
                df_op = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = sheet_name_op)
        else:
            df_op = pd.DataFrame(columns=['SA PartNumber', 'SA Description', 'Total Cost', 'Program Matrix']) 
            sheet_op_missing.append(paths_cpct)
        # append
        df_op['Path'] = paths_cpct
        op_base = op_base.append(df_op, ignore_index = True) 
        op_base = op_base[['SA PartNumber', 'SA Description', 'Total Cost', 'Path']]
        wb.close() # close file
        app.quit() # close app



    #### save file
        # Create target Directory if it doesn't exist
    dirName = '/'.join(fileName_busa[0].split('/')[:-2])+'/CPCT_Consolidation'
    if not os.path.exists(dirName):
        os.mkdir(dirName) 

        #create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(dirName+'/CPCT_Consolidation_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx', engine='xlsxwriter')

        #write each DataFrame to a specific sheet
    busa_base.to_excel(writer, sheet_name='BU SA', index=False)
    op_base.to_excel(writer, sheet_name='OP', index=False)

    #format header
        # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['BU SA']
    worksheet2 = writer.sheets['OP']

        # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#8EA9DB'})

    worksheet.set_column("A:E", 25)
    worksheet2.set_column("A:E", 25)

        # Write the column headers with the defined format.
    for col_num, value in enumerate(busa_base.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for col_num, value in enumerate(op_base.columns.values):
        worksheet2.write(0, col_num, value, header_format)

        #close the Pandas Excel writer and output the Excel file
    writer.save()
    writer.close()
    base.range('B19').value = dirName+'/CPCT_Consolidation_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx'
    base.range('C29', 'Q200').clear_contents()
    base.range('C29').value = pd.DataFrame(fileName_busa)
    #revised by Lily (lily.chen1@hp.com) to prevent 'OptionSA_SUM' sheet missing on 2022/06/10
    #print  CPCT that op sheet missing in the Quote tool
    if len(sheet_op_missing) > 0:
        base.range('S29').value = pd.DataFrame(sheet_op_missing)
        #pop up warning
        MissingOP2()
    #Let user user know the code is finished (pop up Done buttion.)(Please refer to function1-3)
    finish_code()

## 4-4.function4-1: Let user chooses the CPCT files ---Start-------------------------------------------------------------------------
## build CPCT user interaction form(user enters CPCT file's path) - for cpct_consolidation()
class ConsolidationForm_BUSA(QWidget):
    def __init__(self, name = 'ConsolidationForm_BUSA'):
        super(ConsolidationForm_BUSA,self).__init__()
        self.setWindowTitle(name)
        self.resize(400,150)   # set the pop up widget's size

        # btn 1
        self.btn_chooseFile1 = QPushButton(self)  
        self.btn_chooseFile1.setObjectName("btn_chooseFile")  
        self.btn_chooseFile1.setText("[---Choose CPCT file's Path---]")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_chooseFile1)

        self.setLayout(layout)

        # set the widget's signal
        self.btn_chooseFile1.clicked.connect(self.slot_btn_chooseFile1)
    
    def slot_btn_chooseFile1(self):
        global fileName_busa
        fileName_busa, filetype = QFileDialog.getOpenFileNames(self,  
                                    "Choose CPCT file's Path",  
                                    "", # start path
                                    "Excel File (*.xlsx *.xls *.xlsb);;All Files (*)")   

        if fileName_busa != "":
            self.close()
            


def Consolidation_BUSA():
    app = QApplication(sys.argv)
    consolidationForm_BUSA = ConsolidationForm_BUSA('Quote Consolidation_BUSA')
    consolidationForm_BUSA.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')   

## 4-1.function4-1: Let user chooses the CPCT files ---Done-------------------------------------------------------------------------

#revised by Lily (lily.chen1@hp.com) to prevent 'OptionSA_SUM' sheet missing on 2022/06/10
## 4-2.function4-2: Let user know "OP SUM' sheet is missing in CPCT ---Start----------------------------------------
class MissingOPForm2(QWidget):
    def __init__(self, name = 'MissingOPForm2'):
        super(MissingOPForm2,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_done")  
        self.btn_done.setText("Please check your CPCT: OP SUM sheet is missing")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def MissingOP2():
    app = QApplication(sys.argv)
    missingOPForm2 = MissingOPForm2('CPCT Consolidation')
    missingOPForm2.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 4-2.function4-2: Let user know "OP SUM' sheet is missing in CPCT ---Start----------------------------------------



## 4.function4: CPCT BU SA & Option Consolidation ---Done-----------------------------------------------------------------------------------------------------------------------------------------------------------------



## 5.function5: HP Price Consolidation ---Start-----------------------------------------------------------------------------------------------------------------------------------------------------------------
#(function 5 was added by lily.chen1@hp.com 05/26/2022)
def HP_Price_Consolidation():
    import xlwings as xw

    ## -----below quote is for testing
    # simulation = 'C:/Users/CheLily/OneDrive - HP Inc/Desktop/Quote/Quote Validation-IEC/test-quote tool/Quote Tool_20220308_v7.xlsm' #for test
    # wb = xw.Book(simulation) #for test
    # base = wb.sheets['HP Price Consolidation'] #for test
    ## -----
    base = xw.Book.caller().sheets['HP Price Consolidation'] 
    ## Let user chooses the CPCT, so get "fileName_busa" (Please refer to function2-2)
    Consolidation_BUSA()
    ## Let user enters the Sell Price&CKIT file, so get "fileName_choose1", "fileName_choose2", "fileName_choose3" (Please refer to function2-3)
    choose_files2()

    #### Get BUSA & OP quote from CPCT's BUSA sheet-----------------------------------------------------------------------------------------
    #Add OP Quote -- Lily 2022/05/18
    busa_base = pd.DataFrame(columns=['SA \nLevel 3', 'Description', 'Total Base Unit Cost excluded B/S', 'Path'])
    op_base = pd.DataFrame(columns=['SA PartNumber', 'SA Description', 'Total Cost', 'Path'])
    for paths_cpct in fileName_busa:
        base.range('E22').value = paths_cpct
        #open file
        import xlwings as xw
        app = xw.App(visible=False, add_book=False) # don't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_cpct, update_links=False, read_only=True, ignore_read_only_recommended=True)

        if 'xlsb' in paths_cpct:
            cpct = pd.ExcelFile(paths_cpct, engine='pyxlsb')
        else:
            cpct = pd.ExcelFile(paths_cpct)            
        sheet_names = cpct.sheet_names 

        sheet_name_busa = []
        sheet_name_op = []
        for i in sheet_names:
            if 'BUSA' in i: 
                sheet_name_busa = i            
            if 'BU SA' in i: 
                sheet_name_busa = i 
            if 'OptionSA_SUM' in i: 
                sheet_name_op = i
                
        #load the data
        if 'xlsb' in paths_cpct:
            df_busa = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_busa)
            df_op = pd.read_excel(paths_cpct, engine='pyxlsb', skiprows = 0, sheet_name = sheet_name_op)
        else: 
            df_busa = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = sheet_name_busa)
            df_op = pd.read_excel(paths_cpct, skiprows = 0, sheet_name = sheet_name_op)
    #     df_busa = df_busa.iloc[:,:3].reindex(columns=['SA', 'Description', 'Total Base Unit Cost excluded B/S'])
        df_busa['Path'] = paths_cpct
        df_op['Path'] = paths_cpct
        # append
        busa_base = busa_base.append(df_busa) 
        op_base = op_base.append(df_op) 


        wb.close() # close file
        app.quit() # close app


    #### Get sell price from Sell price file
    if 'xlsb' in fileName_choose1:
        sell_base = pd.read_excel(fileName_choose1, engine='pyxlsb', skiprows = 0, sheet_name = 'Master Data')
    else: 
        sell_base = pd.read_excel(fileName_choose1, skiprows = 0, sheet_name = 'Master Data')
    sell_base['Path'] = fileName_choose1


    #### Get ckit price from CKIT file  ----------------------------------------------------------------------------------------------------
    if 'xlsb' in fileName_choose2:
        df_ckit_nb1 = pd.read_excel(fileName_choose2, engine='pyxlsb', skiprows = 0, sheet_name = 'Doc Kit SKU Summary')
        df_ckit_nb2 = pd.read_excel(fileName_choose2, engine='pyxlsb', skiprows = 0, sheet_name = 'Doc KIT SKU summary for HP')
    else:
        df_ckit_nb1 = pd.read_excel(fileName_choose2, skiprows = 0, sheet_name = 'Doc Kit SKU Summary')
        df_ckit_nb2 = pd.read_excel(fileName_choose2, skiprows = 0, sheet_name = 'Doc KIT SKU summary for HP')

    df_ckit_me = pd.DataFrame(columns = ['HP P/N', 'Description', 'Current Price']) 
    ckit_file = load_workbook(fileName_choose3)
    skit_sheet_names = ckit_file.get_sheet_names()
    for i, j in zip(ckit_file, ckit_file.sheetnames):
        if i.sheet_state == "visible":
            if 'xlsb' in fileName_choose3:
                df_ckit_me1 = pd.read_excel(fileName_choose3, engine='pyxlsb', skiprows = 0, sheet_name = j)
                df_ckit_me = pd.concat((df_ckit_me, df_ckit_me1[['HP P/N', 'Description', 'Current Price']]), axis=0, ignore_index=True)
            else:
                df_ckit_me1 = pd.read_excel(fileName_choose3, skiprows = 0, sheet_name = j) 
                df_ckit_me = pd.concat((df_ckit_me, df_ckit_me1[['HP P/N', 'Description', 'Current Price']]), axis=0, ignore_index=True)

    df_ckit_nb = pd.concat([df_ckit_nb1[['HP P/N', 'Description', 'Current Price']], df_ckit_nb2[['HP P/N', 'Description', 'Current Price']]],axis=0, ignore_index=True)
    ckit_base = pd.concat([df_ckit_nb, df_ckit_me],axis=0, ignore_index=True)    
    ckit_base['Path'] = fileName_choose2

    #clean ckit data - deal with the issue caused by 'Merge cells'--------------------------------------------------------------------------
    ckit_base['HP P/N'] = ckit_base['HP P/N'].str.replace('\n','')
    ckit_base['Current Price'].fillna(method='ffill', inplace=True)

    # Consolidate HP price----------------------------------------------------------------------------------------------------------------------------------------------------
    ckit_base['Category'] = 'CKIT'
    busa_base['Category'] = 'BU SA'
    op_base['Category'] = 'OP'
    sell_base['Category'] = 'Sell Price'
    busa_base = busa_base[['SA \nLevel 3', 'Description', 'Total Base Unit Cost excluded B/S', 'Category', 'Path']]
    op_base = op_base[['SA PartNumber', 'SA Description', 'Total Cost', 'Category', 'Path']]
    busa_base.set_axis(['HP P/N', 'Description', 'Current Price', 'Category', 'Path'],axis='columns', inplace=True)
    op_base.set_axis(['HP P/N', 'Description', 'Current Price', 'Category', 'Path'],axis='columns', inplace=True)
    HP_price_base = pd.concat([ckit_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']], busa_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']], sell_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']], op_base[['HP P/N', 'Description', 'Current Price', 'Category', 'Path']]], axis=0, ignore_index=True)    
   
    #### save file----------------------------------------------------------------------------------------------------------------------------------------------------
        # Create target Directory if it doesn't exist
    dirName = '/'.join(fileName_busa[0].split('/')[:-2])+'/HP Price_Consolidation'
    if not os.path.exists(dirName):
        os.mkdir(dirName) 

        #create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(dirName+'/HP Price_Consolidation_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx', engine='xlsxwriter')

        #write each DataFrame to a specific sheet
    HP_price_base.to_excel(writer, sheet_name='HP Price', index=False)
    #format header
        # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['HP Price']

        # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#8EA9DB'})

    worksheet.set_column("A:E", 25)

        # Write the column headers with the defined format.
    for col_num, value in enumerate(HP_price_base.columns.values):
        worksheet.write(0, col_num, value, header_format)
        
        #close the Pandas Excel writer and output the Excel file--------------------------------------------------------------------------
    writer.save()
    writer.close()
    base.range('B19').value = dirName+'/HP Price_Consolidation_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx'
    base.range('C29', 'Q200').clear_contents()
    base.range('C29').value = pd.DataFrame(fileName_busa)
    #Let user user know the code is finished (pop up Done buttion.)(Please refer to function1-3)
    finish_code()

## 5.function5: HP Price Consolidation ---End-----------------------------------------------------------------------------------------------------------------------------------------------------------------

## 6.function6: TW CM Quote Consolidation---Start-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

## 6-1.function6-1: Pop out warning---End-------------------------------------------------------------------------

def TW_quote_consolidation():
    #Let user enters effectively date(Please reder to function3-2)
    consolidation_getdateForm()    
    #Let user chooses the quote file to consolidate(Please reder to function3-1)
    app = QApplication(sys.argv)
    consolidationForm = ConsolidationForm('Quote Consolidation')
    consolidationForm.show()
    
    base = xw.Book.caller().sheets['TW Quote Consolidation']
    
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
    
    
    new_effective_day = 'Effective '+ text
    
    df = pd.DataFrame(columns=['Part', 'Description', new_effective_day, 'Platform', 'ODM'])
    df2 = pd.DataFrame(columns=['SKU', 'AV', 'ODM', 'Path'])
    test_platform = []
    for paths_quote in fileName_choose:
        base.range('E22').value = paths_quote  
        test_platform.append(paths_quote)
#         print(paths_quote)
         #get quote_program_matrix sheet's name
        if 'xlsb' in paths_quote:
            quote_program_matrix = pd.ExcelFile(paths_quote, engine='pyxlsb')
        else:
            quote_program_matrix = pd.ExcelFile(paths_quote)

        sheet_names = quote_program_matrix.sheet_names 

        app = xw.App(visible=False, add_book=False) # can't see the file
        app.display_alerts = False # close alert
        app.screen_updating = False # close screen update
        wb = app.books.open(paths_quote, update_links=False, read_only=True, ignore_read_only_recommended=True)

        quote_file_name = paths_quote.split("/")[-1]

        #load the data-----------------------------------------------------------------------------------------------------------
        if 'Summary' in sheet_names:
            if 'xlsb' in paths_quote:
                df_twquote = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = 'Summary')
            else: 
                df_twquote = pd.read_excel(paths_quote, skiprows = 0, sheet_name = 'Summary')
        else:
            df_twquote = pd.DataFrame(columns=['Part', 'Description', new_effective_day, 'Platform', 'ODM'])
        if 'SKU' in sheet_names:
            if 'xlsb' in paths_quote:
                df_twquote_BOM = pd.read_excel(paths_quote, engine='pyxlsb', skiprows = 0, sheet_name = 'SKU')
            else:
                df_twquote_BOM = pd.read_excel(paths_quote, skiprows = 0, sheet_name = 'SKU')
        else:
            df_twquote_BOM = pd.DataFrame(columns=['SKU', 'AV', 'ODM', 'Path'])
            
        # get "Summary" & "SKU" sheets' data--------------------------------------------------------------------------
        ## "Summary" 
        df_twquote = df_twquote.iloc[:,[1,3,8,10]]
        odm = ''
        if 'Compal' in paths_quote:
            odm = 'Compal'
        elif 'compal' in paths_quote:
            odm = 'Compal'
        elif 'Foxconn' in paths_quote:
            odm = 'Foxconn'
        elif 'foxconn' in paths_quote:
            odm = 'Foxconn'           
        elif 'Pegatron' in paths_quote:
            odm = 'Pegatron'
        elif 'pegatron' in paths_quote:
            odm = 'Pegatron'            
        elif 'Quanta' in paths_quote:
            odm = 'Quanta'
        elif 'quanta' in paths_quote:
            odm = 'Quanta'
        elif 'Inventec' in  paths_quote:
            odm = 'Inventc'  
        elif 'inventec' in  paths_quote:
            odm = 'Inventc' 
        else:
            odm = 'N/A'
        df_twquote.loc[:, 'ODM'] = odm 
        df_twquote.set_axis(['Part', 'Description', new_effective_day, 'Platform', 'ODM'], axis='columns', inplace=True)
        df = df.append(df_twquote) 
        
        df_twquote_BOM = df_twquote_BOM.iloc[:, [1,2]]
        df_twquote_BOM = df_twquote_BOM.dropna(how='all')
        #df_twquote_BOM.loc[:, 'ODM'] = odm 
        df_twquote_BOM['ODM'] = odm 
        #df_twquote_BOM.loc[:, 'Path'] = paths_quote
        df_twquote_BOM['Path'] = paths_quote
        df_twquote_BOM.set_axis(['SKU', 'AV', 'ODM', 'Path'], axis='columns', inplace=True)
        df2 = df2.append(df_twquote_BOM)        
        wb.close() # close file
        app.quit() # close app

    # seperate AV & SKU from df
    df = df.dropna(how='all', subset=['Part'])
    twquote_avlist = [i for i in df['Part'] if 'AV' in i]
    df_twquote_AV = df[df.iloc[:,0].isin(twquote_avlist)]
    df_twquote_SKU = df[~df.iloc[:,0].isin(twquote_avlist)]

    ## check if 'SKU' has repetitive "SKU" + "AV"?      
    df_twquote_BOM_repe = df2[df2.duplicated(subset=df_twquote_BOM.columns)]

        
    #save file-------------------------------------------------------------------------------------------------------------------------------------
        # Create target Directory if it doesn't exist
    dirName = '/'.join(fileName_choose[0].split('/')[:-2])+'/TW Quote Consolidation_' + text 
    if not os.path.exists(dirName):
        os.mkdir(dirName) 

        #create a Pandas Excel writer using XlsxWriter as the engine
    writer = pd.ExcelWriter(dirName+'/TW Quote Consolidation_'+text+'_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx', engine='xlsxwriter')

        #write each DataFrame to a specific sheet
    df_twquote_AV.to_excel(writer, sheet_name='AV', index=False)
    df_twquote_SKU.to_excel(writer, sheet_name='SKU', index=False)
    #format header
        # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['AV']
    worksheet2 = writer.sheets['SKU']

        # Add a header format.
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#8EA9DB'})

    worksheet.set_column("A:E", 25)
    worksheet2.set_column("A:E", 25)

        # Write the column headers with the defined format.
    for col_num, value in enumerate(df_twquote_AV.columns.values):
        worksheet.write(0, col_num, value, header_format)
    for col_num, value in enumerate(df_twquote_SKU.columns.values):
        worksheet2.write(0, col_num, value, header_format)


        #close the Pandas Excel writer and output the Excel file
    writer.save()
    writer.close()
    #print save file's path
    base.range('B19').value = dirName+'/TW Quote Consolidation_'+text+'_'+datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')+'.xlsx'
    df3 = pd.DataFrame(test_platform)
    base.range('C29', 'O311').clear_contents()
    base.range('C29').value = df3
    
    #if 'SKU' has repetitive "SKU" + "AV", show in the sheet
    if len(df_twquote_BOM_repe) != 0:
        base.range('C44').value = 'TW CM Quote has repetition SKU+AV as below'       
        base.range('C45').value = df_twquote_BOM_repe
    #pop up alert
        bom_warn_code()     
    finish_code()

#bom_warn_code
class BOM_Warn_Form(QWidget):
    def __init__(self, name = 'BOM_Warn_Form'):
        super(BOM_Warn_Form,self).__init__()
        self.setWindowTitle(name)
        self.resize(300,150)   # set the pop up widget's size

        # btn 1
        self.btn_done = QPushButton(self)  
        self.btn_done.setObjectName("btn_tbd_warn")  
        self.btn_done.setText("Please check your 'SKU' sheet: it has repetition SKU+AV")
        
        # set the widget's layout
        layout = QVBoxLayout()
        layout.addWidget(self.btn_done)


        self.setLayout(layout)


        # set the widget's signal
        self.btn_done.clicked.connect(self.close)
        
def bom_warn_code():
    app = QApplication(sys.argv)
    bom_Warn_Form = BOM_Warn_Form('Quote Validation')
    bom_Warn_Form.show()
    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing Window...')
## 6-1.function6-1: Pop out warning---End-------------------------------------------------------------------------
        
## 6.function6: TW CM Quote Consolidation---End-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
